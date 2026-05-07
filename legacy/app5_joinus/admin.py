from django.contrib import admin
from django.http import HttpResponse
import csv
from .models import Application
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

@admin.register(Application)
class ApplicationAdmin(admin.ModelAdmin):
    list_display = (
        'user',            
        'fullname',         
        'email',
        'phonenumber',
        'kakaotalkID',
        'gender',
        'academicyear',
        'intendedmajor',
        'labpreference',
        'created_at',
    )
    list_display_links = ('user', 'fullname')  
    list_filter = ('labpreference', 'created_at',)
    search_fields = (
        'user__username', 'user__email',
        'user__first_name', 'user__last_name',
        'motivation',
    )
    date_hierarchy = 'created_at'
    ordering = ('-created_at',)
    list_select_related = ('user',) 
    readonly_fields = (
        'fullname', 'email', 'phonenumber', 'kakaotalkID',
        'gender', 'academicyear', 'intendedmajor',
        'labpreference', 'motivation', 'created_at',
    )
    fieldsets = (
        ('User Info', {
            'fields': (
                'fullname', 'email', 'phonenumber', 'kakaotalkID',
                'gender', 'academicyear', 'intendedmajor',
            )
        }),
        ('Application Info', {
            'fields': ('labpreference', 'motivation', 'created_at')
        }),
    )

    actions = ['export_as_csv','export_as_xlsx']

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.select_related('user', 'user__profile')

    def export_as_csv(self, request, queryset):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=applications.csv'
        writer = csv.writer(response)
        writer.writerow([
            'username', 'fullname', 'email', 'phonenumber', 'kakaotalkID',
            'academicyear', 'intendedmajor', 'gender',
            'labpreference', 'motivation', 'created_at'
        ])
        for obj in queryset:
            writer.writerow([
                obj.user.username,
                obj.fullname(),          
                obj.email(),
                obj.phonenumber(),
                obj.kakaotalkID(),
                obj.academicyear(),
                obj.intendedmajor(),
                obj.gender(),
                obj.labpreference,
                obj.motivation,
                obj.created_at,
            ])
        return response
    export_as_csv.short_description = "Export selected to CSV"




    def export_as_xlsx(self, request, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "Applications"

        headers = [
            'username', 'fullname', 'email', 'phonenumber', 'kakaotalkID',
            'academicyear', 'intendedmajor', 'gender',
            'labpreference', 'motivation', 'created_at'
        ]
        ws.append(headers)

        for obj in queryset.select_related('user', 'user__profile'):
            ws.append([
                obj.user.username,
                obj.fullname(),
                obj.email(),
                obj.phonenumber(),
                obj.kakaotalkID(),
                obj.academicyear(),
                obj.intendedmajor(),
                obj.gender(),
                obj.labpreference,
                obj.motivation,
                obj.created_at.strftime('%Y-%m-%d %H:%M'),
            ])

        for col_idx, _ in enumerate(headers, start=1):
            max_len = 0
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                cell_val = row[0].value
                cell_len = len(str(cell_val)) if cell_val is not None else 0
                max_len = max(max_len, cell_len)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        resp['Content-Disposition'] = 'attachment; filename=applications.xlsx'
        return resp
    export_as_xlsx.short_description = "Export selected to Excel (.xlsx)"